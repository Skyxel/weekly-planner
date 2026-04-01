# weekly_planner/excel_export.py

import io
from typing import List, Optional

import numpy as np
import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .models import PlannerConfig, PlanResult

# Colori
_HEADER_BG = "1F4E79"      # Blu scuro per intestazioni giorni
_HOUR_BG = "2E75B6"        # Blu medio per colonna ore
_OCCUPIED_BG = "DEEAF1"    # Azzurro chiaro per celle occupate
_EMPTY_BG = "F8F9FA"       # Grigio molto chiaro per celle vuote
_TITLE_BG = "1F4E79"       # Stesso del header per riga titolo

_WHITE = "FFFFFF"
_DARK = "1A1A1A"

_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_THICK_BOTTOM = Side(style="medium", color="1F4E79")
_HEADER_BORDER = Border(
    left=_THIN, right=_THIN,
    top=_THIN, bottom=Side(style="medium", color="1F4E79")
)

DAY_LABELS = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]


def _ensure_names(length: int, custom: Optional[List[str]], prefix: str) -> List[str]:
    if custom is not None and len(custom) == length:
        return custom
    return [f"{prefix} {i + 1}" for i in range(length)]


def _header_cell(ws, row: int, col: int, value: str, bg: str = _HEADER_BG) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=_WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _HEADER_BORDER


def _hour_label_cell(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=_WHITE, size=9)
    cell.fill = PatternFill("solid", fgColor=_HOUR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDER


def _data_cell(ws, row: int, col: int, value: str, occupied: bool) -> None:
    cell = ws.cell(row=row, column=col, value=value if occupied else "")
    if occupied:
        cell.font = Font(size=9, color=_DARK)
        cell.fill = PatternFill("solid", fgColor=_OCCUPIED_BG)
    else:
        cell.font = Font(size=9, color="AAAAAA")
        cell.fill = PatternFill("solid", fgColor=_EMPTY_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER


def _write_timetable_sheet(
    ws,
    P: np.ndarray,           # shape (days, daily_hours, num_entities) — vedi sotto
    days: int,
    daily_hours: int,
    day_labels: List[str],
    hour_labels: List[str],
    entity_index: int,       # quale classe o professore stiamo scrivendo
    cell_fn,                 # callable(P, d, h, entity_index) -> (text, is_occupied)
) -> None:
    """
    Scrive un foglio orario nel worksheet `ws`.
    Riga 1: intestazioni giorni.
    Righe 2..(daily_hours+1): label ora + celle dati.
    """
    # Riga intestazione
    _hour_label_cell(ws, 1, 1, "Ora \\ Giorno")
    for d, label in enumerate(day_labels[:days]):
        _header_cell(ws, 1, d + 2, label)

    # Righe dati
    for h in range(daily_hours):
        row = h + 2
        _hour_label_cell(ws, row, 1, hour_labels[h])
        for d in range(days):
            text, occupied = cell_fn(P, d, h, entity_index)
            _data_cell(ws, row, d + 2, text, occupied)

    # Larghezze colonne
    ws.column_dimensions[get_column_letter(1)].width = 14
    for d in range(days):
        ws.column_dimensions[get_column_letter(d + 2)].width = 16

    # Altezze righe
    ws.row_dimensions[1].height = 22
    for h in range(daily_hours):
        ws.row_dimensions[h + 2].height = 20


def render_classes_excel(
    result: PlanResult,
    config: PlannerConfig,
    plan_index: int = 0,
) -> bytes:
    """
    Genera un file Excel con una pagina per ogni classe.
    Ogni pagina mostra ore × giorni con il nome del professore in ogni slot.
    """
    if not result.plans:
        raise ValueError("Nessun piano disponibile per generare l'Excel.")

    P = result.plans[plan_index]
    days = config.days
    daily_hours = config.daily_hours
    m = config.num_classes

    class_names = _ensure_names(m, config.class_names, "Classe")
    professor_names = _ensure_names(config.num_professors, config.professor_names, "Prof")

    day_labels = DAY_LABELS[:days]
    if config.hour_names and len(config.hour_names) == daily_hours:
        hour_labels = config.hour_names
    else:
        hour_labels = [f"Ora {h + 1}" for h in range(daily_hours)]

    week_label = ""
    if result.week_labels and len(result.week_labels) > plan_index:
        week_label = result.week_labels[plan_index]

    def cell_for_class(P, d, h, c):
        prof_id = int(P[d, h, c])
        if prof_id == 0:
            return "", False
        return professor_names[prof_id - 1], True

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # rimuove il foglio vuoto di default

    for c, cname in enumerate(class_names):
        sheet_name = _safe_sheet_name(cname)
        ws = wb.create_sheet(title=sheet_name)

        # Titolo compatto sopra la tabella
        title = f"Piano orario – {cname}"
        if week_label:
            title += f"  ({week_label})"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days + 1)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, color=_WHITE, size=11)
        title_cell.fill = PatternFill("solid", fgColor=_TITLE_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # Sposta la tabella alla riga 2
        _write_timetable_in_ws(ws, P, days, daily_hours, day_labels, hour_labels, c, cell_for_class, start_row=2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def render_professors_excel(
    result: PlanResult,
    config: PlannerConfig,
    plan_index: int = 0,
) -> bytes:
    """
    Genera un file Excel con una pagina per ogni professore.
    Ogni pagina mostra ore × giorni con il nome della classe in ogni slot.
    """
    if not result.plans:
        raise ValueError("Nessun piano disponibile per generare l'Excel.")

    P = result.plans[plan_index]
    days = config.days
    daily_hours = config.daily_hours
    m = config.num_classes
    n = config.num_professors

    class_names = _ensure_names(m, config.class_names, "Classe")
    professor_names = _ensure_names(n, config.professor_names, "Prof")

    day_labels = DAY_LABELS[:days]
    if config.hour_names and len(config.hour_names) == daily_hours:
        hour_labels = config.hour_names
    else:
        hour_labels = [f"Ora {h + 1}" for h in range(daily_hours)]

    week_label = ""
    if result.week_labels and len(result.week_labels) > plan_index:
        week_label = result.week_labels[plan_index]

    def cell_for_prof(P, d, h, p):
        classes_here = [class_names[c] for c in range(m) if int(P[d, h, c]) == p + 1]
        if not classes_here:
            return "", False
        return ", ".join(classes_here), True

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for p, pname in enumerate(professor_names):
        sheet_name = _safe_sheet_name(pname)
        ws = wb.create_sheet(title=sheet_name)

        title = f"Piano orario – {pname}"
        if week_label:
            title += f"  ({week_label})"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days + 1)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, color=_WHITE, size=11)
        title_cell.fill = PatternFill("solid", fgColor=_TITLE_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        _write_timetable_in_ws(ws, P, days, daily_hours, day_labels, hour_labels, p, cell_for_prof, start_row=2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_timetable_in_ws(
    ws,
    P: np.ndarray,
    days: int,
    daily_hours: int,
    day_labels: List[str],
    hour_labels: List[str],
    entity_index: int,
    cell_fn,
    start_row: int = 1,
) -> None:
    """Scrive la griglia oraria nel worksheet a partire da start_row."""
    header_row = start_row

    # Intestazione colonne (giorni)
    _hour_label_cell(ws, header_row, 1, "Ora \\ Giorno")
    for d, label in enumerate(day_labels[:days]):
        _header_cell(ws, header_row, d + 2, label)

    # Righe dati
    for h in range(daily_hours):
        row = header_row + 1 + h
        _hour_label_cell(ws, row, 1, hour_labels[h])
        for d in range(days):
            text, occupied = cell_fn(P, d, h, entity_index)
            _data_cell(ws, row, d + 2, text, occupied)

    # Larghezze
    ws.column_dimensions[get_column_letter(1)].width = 14
    for d in range(days):
        ws.column_dimensions[get_column_letter(d + 2)].width = 16

    # Altezze
    ws.row_dimensions[header_row].height = 22
    for h in range(daily_hours):
        ws.row_dimensions[header_row + 1 + h].height = 20


def _safe_sheet_name(name: str) -> str:
    """Tronca e sanitizza il nome del foglio per Excel (max 31 chars, no caratteri speciali)."""
    forbidden = r"/\?*:[]"
    cleaned = "".join(c if c not in forbidden else "_" for c in name)
    return cleaned[:31]
